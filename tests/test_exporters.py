"""
Тесты для всех экспортеров (Excel, JSON, CSV, HTML).
"""

import pytest
import json
import csv
from pathlib import Path
from excel_export import ExcelExporter
from json_export import JSONExporter
from csv_export import CSVExporter
from html_export import HTMLExporter
from compare import Compare


@pytest.fixture
def sample_comparison_results():
    """Фикстура с примерными результатами сравнения."""
    return [
        {
            "index_1": 1,
            "text_1": "Первый абзац",
            "full_path_1": "Раздел 1",
            "page_1": 1,
            "index_2": 1,
            "text_2": "Первый абзац",
            "full_path_2": "Раздел 1",
            "page_2": 1,
            "status": "identical",
            "similarity": 1.0,
            "differences": [],
            "change_description": "",
            "change_type": "Без изменений",
            "change_subtype": "",
            "llm_response": ""
        },
        {
            "index_1": 2,
            "text_1": "Второй абзац",
            "full_path_1": "Раздел 1",
            "page_1": 1,
            "index_2": 2,
            "text_2": "Второй измененный абзац",
            "full_path_2": "Раздел 1",
            "page_2": 1,
            "status": "modified",
            "similarity": 0.8,
            "differences": ["изменен"],
            "change_description": "Раздел 1, страница 1. 'Второй абзац' изменено на 'Второй измененный абзац'",
            "change_type": "Изменение содержания",
            "change_subtype": "Изменение смысла",
            "llm_response": ""
        },
        {
            "index_1": None,
            "text_1": None,
            "full_path_1": None,
            "page_1": None,
            "index_2": 3,
            "text_2": "Новый абзац",
            "full_path_2": "Раздел 1",
            "page_2": 2,
            "status": "added",
            "similarity": 0.0,
            "differences": [],
            "change_description": "Раздел 1, страница 2. Добавлен абзац: 'Новый абзац'",
            "change_type": "Добавлен",
            "change_subtype": "Добавление абзаца",
            "llm_response": ""
        }
    ]


@pytest.fixture
def sample_statistics():
    """Фикстура с примерной статистикой."""
    return {
        "total": 3,
        "identical": 1,
        "modified": 1,
        "added": 1,
        "deleted": 0,
        "identical_percent": 33.3,
        "modified_percent": 33.3,
        "added_percent": 33.3,
        "deleted_percent": 0.0,
        "change_types": {
            "Без изменений": 1,
            "Изменение содержания": 1,
            "Добавлен": 1
        },
        "tables_total_1": 0,
        "tables_total_2": 0,
        "tables_changed": 0,
        "images_total_1": 0,
        "images_total_2": 0,
        "images_changed": 0,
        "llm_analyzed": 0
    }


class TestExcelExporter:
    """Тесты для Excel экспортера."""
    
    def test_excel_export_basic(self, tmp_path, sample_comparison_results, sample_statistics):
        """Тест базового экспорта в Excel."""
        output_file = tmp_path / "test.xlsx"
        exporter = ExcelExporter(str(output_file))
        
        exporter.export_comparison(
            sample_comparison_results,
            sample_statistics,
            "test1.docx",
            "test2.docx"
        )
        
        assert output_file.exists()
        assert output_file.stat().st_size > 0
    
    def test_excel_export_with_subtype(self, tmp_path, sample_comparison_results, sample_statistics):
        """Тест экспорта в Excel с подтипами изменений."""
        output_file = tmp_path / "test.xlsx"
        exporter = ExcelExporter(str(output_file))
        
        exporter.export_comparison(
            sample_comparison_results,
            sample_statistics,
            "test1.docx",
            "test2.docx"
        )
        
        # Проверяем, что файл создан и содержит данные
        assert output_file.exists()
        
        # Можно добавить проверку содержимого через openpyxl
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        ws = wb.active
        
        # Проверяем наличие столбца "Подтип изменений"
        headers = [cell.value for cell in ws[1]]
        assert "Подтип изменений" in headers


class TestJSONExporter:
    """Тесты для JSON экспортера."""
    
    def test_json_export_pretty(self, tmp_path, sample_comparison_results, sample_statistics):
        """Тест экспорта в JSON с форматированием."""
        output_file = tmp_path / "test.json"
        exporter = JSONExporter(str(output_file), pretty=True)
        
        exporter.export_comparison(
            sample_comparison_results,
            sample_statistics,
            "test1.docx",
            "test2.docx"
        )
        
        assert output_file.exists()
        
        with open(output_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            assert "comparison_results" in data
            assert "statistics" in data
            assert len(data["comparison_results"]) == 3
            # Проверяем наличие change_subtype
            assert "change_subtype" in data["comparison_results"][0]
    
    def test_json_export_compact(self, tmp_path, sample_comparison_results, sample_statistics):
        """Тест экспорта в компактный JSON."""
        output_file = tmp_path / "test.json"
        exporter = JSONExporter(str(output_file), pretty=False)
        
        exporter.export_comparison(
            sample_comparison_results,
            sample_statistics,
            "test1.docx",
            "test2.docx"
        )
        
        assert output_file.exists()
        
        with open(output_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            assert "comparison_results" in data
    
    def test_json_export_with_filters(self, tmp_path, sample_comparison_results, sample_statistics):
        """Тест экспорта в JSON с фильтрами."""
        output_file = tmp_path / "test.json"
        exporter = JSONExporter(str(output_file), pretty=True)
        
        filters = {
            "status": ["modified", "added"]
        }
        
        exporter.export_comparison(
            sample_comparison_results,
            sample_statistics,
            "test1.docx",
            "test2.docx",
            filters=filters
        )
        
        with open(output_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            # Должны остаться только modified и added
            assert len(data["comparison_results"]) == 2
            for result in data["comparison_results"]:
                assert result["status"] in ["modified", "added"]


class TestCSVExporter:
    """Тесты для CSV экспортера."""
    
    def test_csv_export_basic(self, tmp_path, sample_comparison_results, sample_statistics):
        """Тест базового экспорта в CSV."""
        exporter = CSVExporter(str(tmp_path))
        
        exporter.export_comparison(
            sample_comparison_results,
            sample_statistics,
            "test1.docx",
            "test2.docx"
        )
        
        # Проверяем, что созданы CSV файлы
        csv_files = list(tmp_path.glob("*.csv"))
        assert len(csv_files) > 0
        
        # Проверяем содержимое основного файла
        comparison_file = [f for f in csv_files if "comparison" in f.name][0]
        with open(comparison_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            headers = next(reader)
            assert "Подтип изменений" in headers
            
            rows = list(reader)
            assert len(rows) == 3
    
    def test_csv_export_changes_only(self, tmp_path, sample_comparison_results, sample_statistics):
        """Тест экспорта только изменений в CSV."""
        exporter = CSVExporter(str(tmp_path))
        
        exporter.export_comparison(
            sample_comparison_results,
            sample_statistics,
            "test1.docx",
            "test2.docx"
        )
        
        # Проверяем наличие файла с изменениями
        changes_file = tmp_path / "test1_vs_test2_changes_only_*.csv"
        csv_files = list(tmp_path.glob("*changes_only*.csv"))
        if csv_files:
            with open(csv_files[0], 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                headers = next(reader)
                rows = list(reader)
                # Должны быть только изменения (не identical)
                for row in rows:
                    if len(row) > 1:
                        assert row[1] != "identical"


class TestHTMLExporter:
    """Тесты для HTML экспортера."""
    
    def test_html_export_basic(self, tmp_path, sample_comparison_results, sample_statistics):
        """Тест базового экспорта в HTML."""
        output_file = tmp_path / "test.html"
        exporter = HTMLExporter(str(output_file))
        
        exporter.export_comparison(
            sample_comparison_results,
            sample_statistics,
            "test1.docx",
            "test2.docx"
        )
        
        assert output_file.exists()
        
        with open(output_file, 'r', encoding='utf-8') as f:
            content = f.read()
            assert "<html" in content.lower()
            assert "Подтип изменений" in content
            assert "test1.docx" in content
            assert "test2.docx" in content
    
    def test_html_export_structure(self, tmp_path, sample_comparison_results, sample_statistics):
        """Тест структуры HTML экспорта."""
        output_file = tmp_path / "test.html"
        exporter = HTMLExporter(str(output_file))
        
        exporter.export_comparison(
            sample_comparison_results,
            sample_statistics,
            "test1.docx",
            "test2.docx"
        )
        
        with open(output_file, 'r', encoding='utf-8') as f:
            content = f.read()
            # Проверяем наличие основных секций
            assert "Статистика" in content or "статистика" in content.lower()
            assert "<table" in content.lower()
            assert "<style" in content.lower()

