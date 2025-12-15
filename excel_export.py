"""
Модуль для экспорта результатов сравнения в Excel файл.

Класс ExcelExporter предоставляет функционал для:
- Создания структурированных Excel файлов с результатами сравнения
- Форматирования таблиц с цветовой индикацией
- Создания отдельных листов для разных типов данных
- Экспорта статистики и детальных изменений

Создаваемые листы:
1. "Сравнение" - полное сравнение всех абзацев
2. "Только изменения" - фильтрованный список только измененных элементов
3. "Статистика" - общая статистика сравнения
4. "Таблицы" - изменения в таблицах
5. "Изображения" - изменения в изображениях
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict
from datetime import datetime
from config import config
from logger_config import logger
from exceptions import ExportError


class ExcelExporter:
    """
    Класс для экспорта результатов сравнения в Excel.
    
    Создает структурированный Excel файл с:
    - Детальным сравнением абзацев
    - Фильтрованным списком изменений
    - Статистикой сравнения
    - Изменениями в таблицах и изображениях
    - Цветовой индикацией статусов
    """
    
    def __init__(self, output_path: str):
        """
        Инициализация экспортера.
        
        Args:
            output_path: Путь к выходному Excel файлу (будет создан или перезаписан)
        """
        self.output_path = output_path
        self.workbook = Workbook()  # Создание новой рабочей книги
        self.workbook.remove(self.workbook.active)  # Удаление дефолтного листа
    
    def export_comparison(self, comparison_results: List[Dict], 
                         statistics: Dict, file1_name: str, file2_name: str,
                         table_changes: List[Dict] = None,
                         image_changes: List[Dict] = None):
        """
        Экспорт результатов сравнения в Excel.
        
        Args:
            comparison_results: Список результатов сравнения
            statistics: Статистика сравнения
            file1_name: Имя первого файла
            file2_name: Имя второго файла
            table_changes: Список изменений таблиц
            image_changes: Список изменений изображений
        """
        # Создание листа с результатами сравнения
        ws_results = self.workbook.create_sheet("Сравнение", 0)
        self._create_comparison_sheet(ws_results, comparison_results, file1_name, file2_name)
        
        # Создание листа только с изменениями
        changes_only = [r for r in comparison_results if r.get("status") != "identical"]
        if changes_only:
            ws_changes = self.workbook.create_sheet("Только изменения", 1)
            self._create_changes_only_sheet(ws_changes, changes_only, file1_name, file2_name)
        
        # Создание листа со статистикой
        ws_stats = self.workbook.create_sheet("Статистика", 2)
        self._create_statistics_sheet(ws_stats, statistics, file1_name, file2_name)
        
        # Создание листа с изменениями таблиц
        if table_changes:
            ws_tables = self.workbook.create_sheet("Таблицы", 3)
            self._create_tables_sheet(ws_tables, table_changes)
        
        # Создание листа с изменениями изображений
        if image_changes:
            ws_images = self.workbook.create_sheet("Изображения", 4)
            self._create_images_sheet(ws_images, image_changes)
        
        # Сохранение файла
        self.workbook.save(self.output_path)
    
    def _create_comparison_sheet(self, worksheet, comparison_results: List[Dict],
                                 file1_name: str, file2_name: str):
        """Создание листа с результатами сравнения."""
        # Заголовки
        headers = [
            "№",
            "Статус",
            "Тип исправления",
            f"Полный путь ({file1_name})",
            f"Страница ({file1_name})",
            f"Абзац № ({file1_name})",
            f"Текст ({file1_name})",
            f"Полный путь ({file2_name})",
            f"Страница ({file2_name})",
            f"Абзац № ({file2_name})",
            f"Текст ({file2_name})",
            "Схожесть (%)",
            "Различия",
            "Описание изменений",
            "Ответ LLM"
        ]
        
        # Стили
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заполнение заголовков
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Заполнение данных
        for row_idx, result in enumerate(comparison_results, 2):
            status = result["status"]
            
            # Определение цвета строки в зависимости от статуса
            if status == "identical":
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif status == "modified":
                fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif status == "added":
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:  # deleted
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Перевод статуса на русский
            status_ru = {
                "identical": "Идентичен",
                "modified": "Изменен",
                "added": "Добавлен",
                "deleted": "Удален"
            }.get(status, status)
            
            # Данные строки
            row_data = [
                row_idx - 1,  # №
                status_ru,  # Статус
                result.get("change_type", ""),  # Тип исправления
                result.get("full_path_1") or "",  # Полный путь 1
                result.get("page_1") or "",  # Страница 1
                result.get("index_1") or "",  # Абзац № 1
                result.get("text_1") or "",  # Текст 1
                result.get("full_path_2") or "",  # Полный путь 2
                result.get("page_2") or "",  # Страница 2
                result.get("index_2") or "",  # Абзац № 2
                result.get("text_2") or "",  # Текст 2
                f"{result['similarity'] * 100:.1f}%" if result.get("similarity") else "",  # Схожесть
                "\n".join(result.get("differences", []))[:1000],  # Различия (увеличено для полных текстов)
                result.get("change_description", ""),  # Описание изменений
                result.get("llm_response", "")  # Ответ LLM
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = border
        
        # Настройка ширины столбцов
        column_widths = {
            'A': 8,   # №
            'B': 12,  # Статус
            'C': 20,  # Тип исправления
            'D': 40,  # Полный путь 1
            'E': 10,  # Страница 1
            'F': 12,  # Абзац № 1
            'G': 50,  # Текст 1
            'H': 40,  # Полный путь 2
            'I': 10,  # Страница 2
            'J': 12,  # Абзац № 2
            'K': 50,  # Текст 2
            'L': 12,  # Схожесть
            'M': 60,  # Различия
            'N': 60,  # Описание изменений
            'O': 60   # Ответ LLM
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Фиксация первой строки
        worksheet.freeze_panes = 'A2'
    
    def _create_changes_only_sheet(self, worksheet, comparison_results: List[Dict],
                                   file1_name: str, file2_name: str):
        """
        Создание листа только с изменениями (без идентичных элементов).
        
        Фильтрует результаты, оставляя только:
        - Измененные абзацы
        - Добавленные абзацы
        - Удаленные абзацы
        
        Использует тот же формат, что и основной лист "Сравнение",
        но без идентичных элементов для удобства просмотра только изменений.
        
        Args:
            worksheet: Рабочий лист Excel
            comparison_results: Отфильтрованные результаты (только изменения)
            file1_name: Имя первого файла
            file2_name: Имя второго файла
        """
        # Используем тот же метод форматирования, что и для основного листа
        self._create_comparison_sheet(worksheet, comparison_results, file1_name, file2_name)
    
    def _create_statistics_sheet(self, worksheet, statistics: Dict,
                                file1_name: str, file2_name: str):
        """Создание листа со статистикой."""
        # Заголовок
        title_cell = worksheet.cell(row=1, column=1, value="Статистика сравнения документов")
        title_cell.font = Font(bold=True, size=14)
        worksheet.merge_cells('A1:B1')
        
        # Информация о файлах
        worksheet.cell(row=3, column=1, value="Файл 1:").font = Font(bold=True)
        worksheet.cell(row=3, column=2, value=file1_name)
        worksheet.cell(row=4, column=1, value="Файл 2:").font = Font(bold=True)
        worksheet.cell(row=4, column=2, value=file2_name)
        worksheet.cell(row=5, column=1, value="Дата сравнения:").font = Font(bold=True)
        worksheet.cell(row=5, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        # Статистика
        row = 7
        worksheet.cell(row=row, column=1, value="Показатель").font = Font(bold=True)
        worksheet.cell(row=row, column=2, value="Значение").font = Font(bold=True)
        
        stats_data = [
            ("Всего абзацев", statistics.get("total", 0)),
            ("Идентичных", f"{statistics.get('identical', 0)} ({statistics.get('identical_percent', 0):.1f}%)"),
            ("Измененных", f"{statistics.get('modified', 0)} ({statistics.get('modified_percent', 0):.1f}%)"),
            ("Добавленных", f"{statistics.get('added', 0)} ({statistics.get('added_percent', 0):.1f}%)"),
            ("Удаленных", f"{statistics.get('deleted', 0)} ({statistics.get('deleted_percent', 0):.1f}%)"),
            ("", ""),  # Пустая строка для разделения
            ("Таблицы", ""),
            ("  Всего таблиц в документе 1", statistics.get("tables_total_1", "N/A")),
            ("  Всего таблиц в документе 2", statistics.get("tables_total_2", "N/A")),
            ("  Изменено таблиц", statistics.get("tables_changed", 0)),
            ("", ""),  # Пустая строка для разделения
            ("Изображения", ""),
            ("  Всего изображений в документе 1", statistics.get("images_total_1", "N/A")),
            ("  Всего изображений в документе 2", statistics.get("images_total_2", "N/A")),
            ("  Изменено изображений", statistics.get("images_changed", 0)),
            ("", ""),  # Пустая строка для разделения
            ("LLM анализ", ""),
            ("  Проанализировано элементов", statistics.get("llm_analyzed", 0))
        ]
        
        for stat_name, stat_value in stats_data:
            row += 1
            worksheet.cell(row=row, column=1, value=stat_name)
            worksheet.cell(row=row, column=2, value=stat_value)
        
        # Настройка ширины столбцов
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 30
    
    def _create_tables_sheet(self, worksheet, table_changes: List[Dict]):
        """Создание листа с изменениями таблиц."""
        # Заголовки
        headers = ["№", "Статус", "Название таблицы 1", "Название таблицы 2", "Описание", "Описание изменений"]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заполнение заголовков
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Заполнение данных
        for row_idx, change in enumerate(table_changes, 2):
            status = change["status"]
            
            # Определение цвета строки
            if status == "identical":
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif status == "modified":
                fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            status_ru = {
                "identical": "Идентична",
                "modified": "Изменена",
                "added": "Добавлена",
                "deleted": "Удалена"
            }.get(status, status)
            
            # Формируем описание изменений в ячейках
            cell_changes_desc = ""
            if change.get("cell_changes"):
                cell_changes = change["cell_changes"]
                changes_list = []
                for cc in cell_changes[:10]:  # Ограничиваем количество
                    changes_list.append(f"Строка {cc['row']}, столбец {cc['col']}")
                cell_changes_desc = f"Изменения в: {', '.join(changes_list)}"
                if len(cell_changes) > 10:
                    cell_changes_desc += f" и еще {len(cell_changes) - 10}"
            
            row_data = [
                row_idx - 1,
                status_ru,
                change.get("table_1_name") or change.get("table_1_index") or "",
                change.get("table_2_name") or change.get("table_2_index") or "",
                change.get("description", ""),
                change.get("change_description", "") or cell_changes_desc
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = border
        
        # Настройка ширины столбцов
        worksheet.column_dimensions['A'].width = 8
        worksheet.column_dimensions['B'].width = 12
        worksheet.column_dimensions['C'].width = 30
        worksheet.column_dimensions['D'].width = 30
        worksheet.column_dimensions['E'].width = 40
        worksheet.column_dimensions['F'].width = 60
        
        worksheet.freeze_panes = 'A2'
    
    def _create_images_sheet(self, worksheet, image_changes: List[Dict]):
        """Создание листа с изменениями изображений."""
        # Заголовки
        headers = ["№", "Статус", "Название изображения 1", "Название изображения 2", "Описание", "Описание изменений"]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заполнение заголовков
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Заполнение данных
        for row_idx, change in enumerate(image_changes, 2):
            status = change["status"]
            
            # Определение цвета строки
            if status == "identical":
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            status_ru = {
                "identical": "Идентично",
                "added": "Добавлено",
                "deleted": "Удалено"
            }.get(status, status)
            
            row_data = [
                row_idx - 1,
                status_ru,
                change.get("image_1_name") or change.get("image_1_index") or "",
                change.get("image_2_name") or change.get("image_2_index") or "",
                change.get("description", ""),
                change.get("change_description", "")
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = border
        
        # Настройка ширины столбцов
        worksheet.column_dimensions['A'].width = 8
        worksheet.column_dimensions['B'].width = 12
        worksheet.column_dimensions['C'].width = 30
        worksheet.column_dimensions['D'].width = 30
        worksheet.column_dimensions['E'].width = 40
        worksheet.column_dimensions['F'].width = 60
        
        worksheet.freeze_panes = 'A2'

